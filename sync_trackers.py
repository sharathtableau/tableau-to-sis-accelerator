"""
sync_trackers.py -- keep the three director-facing xlsx trackers in sync
with the accelerator's actual state. Run at every task close-out (part of
the standing six-file checklist; weekly_status.py stays the canonical
report -- these xlsx files mirror it for people who live in Excel).

    python sync_trackers.py
"""

import datetime

import openpyxl

TODAY = datetime.date.today().isoformat()

# ---- Implementation Roadmap: action-item statuses (row -> status) ----------
ROADMAP = "Tableau to SiS - Implementation Roadmap.xlsx"
ROADMAP_STATUS = {          # row: (status)  [col G]
    10: "Done",             # 5 filter values incl. top-N (context = later)
    11: "Done",             # 6 table calcs WINDOW_/RANK/INDEX (RUNNING/LOOKUP refuse)
    12: "Done",             # 7 manual + computed sort
    13: "Done",             # 8 number formats / colour model (closed category)
    14: "In Progress",      # 9 joins/blends: flatten + view generator DONE; blends partial
    15: "Done",             # 10 reference / distribution / trend lines
    16: "Done",             # 11 parameters as widgets
    17: "Not Started",      # 12 dashboard actions (reported drops)
    18: "Not Started",      # 13 custom tooltips (user-deprioritized)
    19: "Not Started",      # 14 images / logos / shapes
}

# ---- Progress Tracker: workstream percentages + narrative -------------------
PROGRESS = "Tableau to SiS - Progress Tracker.xlsx"
PROGRESS_PCT = {            # row: (pct, RAG)  [cols D, E]
    8:  (0.92, "G"),        # parsing & extraction (+FCP tags, relationships)
    9:  (0.80, "G"),        # chart-type library (21 kinds)
    10: (0.97, "G"),        # calc translation (audit_calcs live number)
    11: (0.75, "G"),        # data layer (flatten + semantic view generator; dry-run open)
    12: (0.55, "A"),        # interactivity (params, top-N, drill selector)
    13: (0.70, "G"),        # formatting & layout
    14: (0.80, "G"),        # deployment (dry-run DONE: load + SiS deploy verified)
    15: (0.90, "G"),        # quality & tooling (13 regression gates, 3 audits)
}
PROGRESS_DONE = [
    "• Table-calc engine: WINDOW_*/RANK/INDEX → SQL windows; layered hoist for FIXED-in-agg chains.",
    "• Relationship flatten: multi-table extracts join automatically (E-Commerce 46→68%).",
    "• Top-N filters (field + parameter), hierarchies drill selector, device layouts.",
    "• 2024.3 sample pack: two official workbooks at 96%/100%; corpus now 7.",
    "• Silent gaps closed: custom SQL, relative-date filters, log/reversed axes all report.",
]
PROGRESS_NEXT = [
    "• Bins + histogram; context filters (order of operations).",
    "• Live-connection semantic views (non-extract joins/blends).",
    "• Credentialed Snowflake deploy dry-run (write_pandas).",
    "• Edge-case workbook for joins/multi-fact/RLS/parameter actions.",
]

# ---- Status Tracker: milestone percentages + narrative ----------------------
STATUS = "Tableau to SiS - Status Tracker.xlsx"
STATUS_PCT = {              # row: (pct, RAG)  [cols G, H]
    10: (0.85, "G"),        # M3 data layer (semantic views done; dry-run open)
    11: (0.90, "G"),        # M4 dashboard fidelity
    12: (0.97, "G"),        # M5 calc translation incl. table calcs
    13: (0.55, "A"),        # M6 interactivity
    14: (0.80, "G"),        # M7 deployment (dry-run done)
    15: (0.50, "A"),        # M8 accelerator UI + QC report
}
STATUS_DONE = [
    "• Table-calc engine shipped (WINDOW_*/RANK/INDEX → SQL windows).",
    "• Multi-table relationship extracts flatten automatically; E-Commerce 46→68%.",
    "• Top-N, hierarchies drill, device layouts, silent-gap findings shipped.",
    "• Corpus grew to 7 workbooks (two official 2024.3 samples: 96%/100%).",
    "• Regression suite now 13 gates; weekly report self-updating.",
]
STATUS_NEXT = [
    "• Bins/histogram; context filters.",
    "• Live-source semantic views in Snowflake.",
    "• Credentialed SiS deploy dry-run.",
    "• Edge-case workbook (joins/multi-fact/RLS).",
]


def _set(ws, row, col, value):
    ws.cell(row=row, column=col).value = value


def sync_roadmap():
    wb = openpyxl.load_workbook(ROADMAP)
    ws = wb.worksheets[0]
    for row, status in ROADMAP_STATUS.items():
        _set(ws, row, 7, status)
    wb.save(ROADMAP)
    print("synced", ROADMAP)


def sync_progress():
    wb = openpyxl.load_workbook(PROGRESS)
    ws = wb.worksheets[0]
    _set(ws, 5, 3, TODAY)
    for row, (pct, rag) in PROGRESS_PCT.items():
        _set(ws, row, 4, pct)
        _set(ws, row, 5, rag)
    for i, txt in enumerate(PROGRESS_DONE):
        _set(ws, 18 + i, 2, txt)
    for i, txt in enumerate(PROGRESS_NEXT):
        _set(ws, 18 + i, 5, txt)
    wb.save(PROGRESS)
    print("synced", PROGRESS)


def sync_status():
    wb = openpyxl.load_workbook(STATUS)
    ws = wb.worksheets[0]
    _set(ws, 5, 3, TODAY)
    for row, (pct, rag) in STATUS_PCT.items():
        _set(ws, row, 7, pct)
        _set(ws, row, 8, rag)
    for i, txt in enumerate(STATUS_DONE):
        _set(ws, 19 + i, 2, txt)
    for i, txt in enumerate(STATUS_NEXT):
        _set(ws, 19 + i, 7, txt)
    wb.save(STATUS)
    print("synced", STATUS)


if __name__ == "__main__":
    sync_roadmap()
    sync_progress()
    sync_status()
    print("\nEdit the dicts at the top of this file when statuses move, "
          "then re-run. weekly_status.py remains the canonical report.")
